import pandas as pd
import numpy as np
from Config import config
from pandas import ExcelWriter
import random

###########
# Constants
###########
LIFE_SPAN = 240
LOWER_BOUND_MONTH = -12
ONE_YR = 12
DEFAULT_SHEET_NAME = 'input Default Cohort'

class Simulation:
    vars_path = config.var_path
    '''
    Note on the output parameter, the yearly values of customers are the final (max) and all
    the other values are the sum of the previous six or twlve months. 
    '''
    def __init__(self, data_path):
        self.data_path = data_path
        ### Params
        self.current_month = 0
        self.customer_growth = []
        self.single_month_dic = {}
        self.multi_month_dic = {}

        self.group_cohort = {}
        self.financial_report = pd.DataFrame()
        self.init_global_data()
        self.default_cohort = self.extract_data_as_df(DEFAULT_SHEET_NAME)

    def init_global_data(self, vars_path=vars_path):
        dloader = DataLoader(self.data_path, variable_path = vars_path)
        self.customer_growth = dloader.access_customer_growth_number()
        print(self.customer_growth)
        self.single_month_dic = dloader.create_single_month_dict()
        self.multi_month_dic = dloader.create_multiply_month_dict()

    #####################
    # Access input data #
    #####################
    def extract_data_as_df(self, sheet_name):
        month_dataloader = DataLoader(self.data_path)
        return month_dataloader.extract_dataframe_from_sheet(sheet_name)

    def run_for_this_month(self):
        month = self.current_month
        new_cohort = None
        for item in self.multi_month_dic.keys():
            if month in item:
                temp_df = self.extract_data_as_df(self.multi_month_dic[item])
                new_cohort = Cohort(month, temp_df, self.customer_growth[month])
        if new_cohort == None:
            # print(self.single_month_dic.keys())
            # print(self.single_month_dic)
            if month in self.single_month_dic.keys():
                temp_df = self.extract_data_as_df(self.single_month_dic[month])
                new_cohort = Cohort(month, temp_df, self.customer_growth[month])
            else:
                new_cohort = Cohort(month, self.default_cohort, self.customer_growth[month])
        else:
            pass
        self.group_cohort[month] = new_cohort
        self.run_all_cohort()
        self.current_month += 1

    def run_all_cohort(self):
        for month in self.group_cohort:
            self.group_cohort[month].update_one_month()

    def run_life_time(self):
        for _ in range(LIFE_SPAN):
            self.run_for_this_month()

    ###########
    # Output Results
    ###########
    def output_full_report(self, nice=True):
        if nice:
            self.nice_print_report_format()
        fina_report = pd.DataFrame()
        for month in self.group_cohort:
            if fina_report.empty:
                fina_report = pd.concat([fina_report, self.group_cohort[month].output_financial_report()])
            else:
                add_1 = self.group_cohort[month].output_financial_report()
                add_2 = fina_report.copy()
                fina_report = fina_report.add(add_1)
                # self.assert_addition_reports(add_1, add_2, fina_report)
                # self.assert_addition_reports(add_1, add_2, fina_report)
        return fina_report


    def output_to_excel(self):
        output_file = self.output_full_report(False)
        output = output_file.transpose()
        with ExcelWriter(f'{config.save_path}') as writer:
            output.to_excel(writer, sheet_name='Tests')

    def nice_print_report_format(self):
        print('Current simulation month is: ', self.current_month - 1)
        print('\n-----------')
        print('This financial report is \n')

    ########################
    # Tests Output Results #
    ########################
    def output_numb_cohort_to_excel(self, numb, save_path):
        first_month = self.group_cohort[numb].output_financial_report()
        output = first_month.transpose()
        with ExcelWriter(f'{save_path}') as writer:
            output.to_excel(writer, sheet_name='test_month')

    # def assert_addition_reports(self, add_1, add_2, fina_report):
    #     vars = self.test_output_var.copy()
    #     column_numb = random.randint(0, len(vars)-1)
    #     column_rand = vars[column_numb]
    #     month_rand = random.randint(0, self.current_month - 1)
    #     one_rand = add_1.loc[month_rand, column_rand]
    #     second_rand = add_2.loc[month_rand, column_rand]
    #     # Random index or columns variables and Tests them
    #     assert one_rand + second_rand == fina_report.loc[month_rand, column_rand], f'At month {month_rand} and column {column_rand}, the addition is not successful.'


'''
Cohort steps
init cohort from the dataframe (append the output variables to the end of the pandaDataframe.)
Loop through the function with the workflow variables.
Run function is the same logic as before. 
'''
class Cohort:
    def __init__(self, now_month, dataframe, customer_number):
        self.relative_position = 0
        self.workflow_vars_list = []
        self.output_vars_list = []
        self.start_month = now_month
        self.current_month = now_month
        self.customer_number = customer_number
        self.data = dataframe.copy()
        self.financial_report = pd.DataFrame()
        ## Variables and Canvas
        self.init_variables()
        self.processed_data = self.process_raw_data()
        self.init_canvas_before_start_month()


    def init_variables(self):
        dloader = DataLoader(config.data_path, variable_path = config.var_path)
        self.workflow_vars_list = dloader.output_variables(DataLoader.workflow_vars_sheet_name)
        self.output_vars_list = dloader.output_variables(DataLoader.output_vars_sheet_name)

    def init_canvas_before_start_month(self):
        canvas_before_month = pd.DataFrame(0, index=range(LOWER_BOUND_MONTH, self.start_month), columns=list(self.data.index)+self.output_vars_list)
        print('Canvas column length: ', len(canvas_before_month.columns))
        self.financial_report = canvas_before_month.copy()

    def process_raw_data(self):
        full_var_df = self.append_list_output_vars(self.data).copy()
        ret_df = full_var_df.transpose()
        print('Row dataframe has column: ', len((ret_df.columns)))
        return ret_df

    def append_list_output_vars(self, df):
        for vars in self.output_vars_list:
            df = pd.concat([df, pd.Series(name=vars, dtype='float64').to_frame().T])
        return df

    def update_customer_with_retention(self, row):
        if self.current_month == self.start_month:
            row['Start_Customer_Input_var'] = self.customer_number
            row['Number_of_Customer_Output_Cash_flow'] = self.customer_number
        else:
            row['Start_Customer_Input_var'] = self.customer_number
            # print(self.current_month)
            retention_yr = self.financial_report.loc[self.current_month - 1, 'Retention_yearly_Input_var']
            last_customer_numb = self.financial_report.loc[self.current_month - 1, 'Number_of_Customer_Output_Cash_flow']
            retention_mth = retention_yr ** (1/12)
            row['Number_of_Customer_Output_Cash_flow'] = last_customer_numb * retention_mth
        return row

    def inflate_premium(self, row):
        if (self.current_month - self.start_month) % ONE_YR == 0 and self.current_month != self.start_month:
            row['Start_avg_premium_Input_var'] *= row['Inflation_Input_var']
        else:
            pass
        return row

    def update_one_month(self):
        current_row = self.access_current_month_row()
        current_row['Month_Input_var'] = self.current_month
        current_row = self.update_customer_with_retention(current_row)
        current_row = self.inflate_premium(current_row)
        wc_timeline = self.current_month - ONE_YR
        new_row = self.update_output_var(current_row, wc_timeline)
        self.append_this_month_to_report(new_row)
        ## loop
        self.current_month += 1
        self.relative_position += 1

    def access_current_month_row(self):
        row = self.processed_data.loc[[self.relative_position]]
        row.index = [self.current_month]
        new_row = row.loc[self.current_month]
        return new_row.copy()

    def append_this_month_to_report(self, row):
        assert len(row) == len(self.financial_report.columns), f'The columns number is not consistent.row is {row.index}, report is {(self.financial_report.columns)}'
        row_to_concat = pd.DataFrame(row).transpose()
        self.financial_report = pd.concat([self.financial_report, row_to_concat])

    def update_output_var(self, row, timeline):
        work_cap_timeline = timeline
        for para in self.workflow_vars_list:
            if para == 'Transacted_premium_volume_Output_Profit_Loss' or para == 'Transacted_premium_volume_Output_Cash_flow':
                row[para] = row['Start_avg_premium_Input_var'] * row['Number_of_Customer_Output_Cash_flow']
                # print(f'THis month {self.current_month} premium {row["Start_avg_premium_Input_var"]} customer {row["Number_of_Customer_Output_Cash_flow"]}')
            elif para == 'ow_Origination_Output_Profit_Loss' or para == 'ow_Origination_Output_Cash_flow':
                ratio = self.get_first_or_second_yr_ratio(row['Revenue_share_of_premium_for_new_business_Input_var'], row['Revenue_share_of_premium_for_renewal_Input_var'])
                row[para] = row['Transacted_premium_volume_Output_Profit_Loss'] * ratio
                row['Network_Output_Profit_Loss'] = row[para]
                row['Network_Output_Cash_flow'] = row[para]
            elif para == 'ow_Underwriting_engine_Output_Profit_Loss' or para == 'ow_Underwriting_engine_Output_Cash_flow':
                ratio = self.get_first_or_second_yr_ratio(0, row['Underwriting_Relative_to_premium_based_on_improvement_first_yr_Input_var'])
                row[para] = row['Transacted_premium_volume_Output_Profit_Loss'] * ratio
            elif para == 'ow_Back_office_app_Output_Profit_Loss' or para == 'ow_Back_office_app_Output_Cash_flow':
                ratio = self.get_first_or_second_yr_ratio(0, row['Backoffice_Relative_to_premium_based_on_improvement_first_year_Input_var'])
                row[para] = row['Transacted_premium_volume_Output_Profit_Loss'] * ratio
                # print(f'ow_Back_office_app_Output_Profit_Loss {row[para]} and ratio is {ratio} and transatected {row["Transacted_premium_volume_Output_Profit_Loss"]}')
            elif para == 'Platform_Output_Profit_Loss' or para == 'Platform_Output_Cash_flow':
                assert not np.isnan(row['ow_Back_office_app_Output_Profit_Loss']), f'back office not exist first.'
                assert not np.isnan(row['ow_Underwriting_engine_Output_Profit_Loss']), f'uw engine not exist first.'
                row[para] = row['ow_Back_office_app_Output_Profit_Loss'] + row['ow_Underwriting_engine_Output_Profit_Loss']
            elif para == 'Revenue_Output_Profit_Loss':
                row[para] = row['Platform_Output_Profit_Loss'] + row['Network_Output_Profit_Loss']
            elif para == 'Revenue_Output_Cash_flow':
                row[para] = row['Platform_Output_Cash_flow'] + row['Network_Output_Cash_flow']
            elif para == 'ow_Loss_Output_Profit_Loss' or para == 'ow_Loss_Output_Cash_flow':
                row[para] = 0
            elif para == 'ow_Distribution_channel_Output_Profit_Loss':
                ratio = self.get_first_or_second_yr_ratio(row['Distribution_channel_cost_as_share_of_premium_first_year_Input_var'], row['Distribution_channel_cost_as_share_of_premium_next_year_Input_var'])
                row[para] = ratio * row['Transacted_premium_volume_Output_Profit_Loss']
            elif para == 'ow_Distribution_channel_Output_Cash_flow':
                assert not np.isnan(row['ow_Distribution_channel_Output_Profit_Loss']), f'Profit loss distribution first.'
                ratio = 1 - row['Working_capital_ratio_Distribution_channel_Input_var']
                row[para] = ratio * row['ow_Distribution_channel_Output_Profit_Loss']
            elif para == 'ow_expenses_Output_Profit_Loss':
                ratio = self.get_first_or_second_yr_ratio(row['MGA_expense_ratio_as_share_of_premium_volume_first_year_Input_var'], row['MGA_expense_ratio_as_share_of_premium_volume_next_year_Input_var'])
                row[para] = row['Transacted_premium_volume_Output_Profit_Loss'] * ratio
            elif para == 'ow_expenses_Output_Cash_flow':
                assert not np.isnan(row['ow_expenses_Output_Profit_Loss']), f'To get profit&loss expense first.'
                ratio = 1 - row['Working_capital_ratio_Expenses_Input_var']
                row[para] = row['ow_expenses_Output_Profit_Loss'] * ratio
            elif para == 'ow_outsourcing_Output_Profit_Loss' or para == 'ow_outsourcing_Output_Cash_flow':
                ratio = self.get_first_or_second_yr_ratio(row['MGA_outsourcing_cost_ratio_as_share_of_premium_volume_first_year_Input_var'], row['MGA_outsourcing_cost_ratio_as_share_of_premium_volume_next_year_Input_var'])
                row[para] = ratio * row['Transacted_premium_volume_Output_Profit_Loss']
            elif para == 'Costs_Output_Profit_Loss':
                sum = row['ow_Loss_Output_Profit_Loss'] + row['ow_Distribution_channel_Output_Profit_Loss']
                row[para] = sum + row['ow_expenses_Output_Profit_Loss'] + row['ow_outsourcing_Output_Profit_Loss']
            elif para == 'Costs_Output_Cash_flow':
                sum = row['ow_Loss_Output_Cash_flow'] + row['ow_Distribution_channel_Output_Cash_flow']
                row[para] = sum + row['ow_expenses_Output_Cash_flow'] + row['ow_outsourcing_Output_Cash_flow']
            elif para == 'Net_income_Output_Profit_Loss':
                row[para] = row['Revenue_Output_Profit_Loss'] - row['Costs_Output_Profit_Loss']
            elif para == 'Net_income_Output_Cash_flow':
                row[para] = row['Revenue_Output_Cash_flow'] - row['Costs_Output_Cash_flow']
            elif para == 'Taxes_Output_Profit_Loss':
                row[para] = row['Taxes_as_share_of_net_income_Input_var'] * row['Net_income_Output_Profit_Loss']
            elif para == 'Taxes_Output_Cash_flow':
                row[para] = row['Taxes_Output_Profit_Loss']
            elif para == 'NOPAT_Output_Profit_Loss':
                row['NOPAT_Output_Profit_Loss'] = row['Net_income_Output_Profit_Loss'] - row['Taxes_Output_Profit_Loss']
            elif para == 'NOPAT_Output_Cash_flow':
                if self.current_month == self.start_month:
                    self.financial_report.loc[:self.start_month - 1, 'NOPAT_Output_Cash_flow'] = 0
                row['NOPAT_Output_Cash_flow'] = row['Net_income_Output_Cash_flow'] - row['Taxes_Output_Cash_flow']
            elif para == 'Accumulated_Output_Profit_Loss':
                last_accumulate = 0
                if self.current_month > self.start_month:
                    last_accumulate = self.financial_report.loc[:(self.current_month - 1), 'NOPAT_Output_Profit_Loss'].sum()
                row[para] = last_accumulate + row['NOPAT_Output_Profit_Loss']
            elif para == 'Loss_Output_Profit_Loss_Carrier':
                row[para] = row['Transacted_premium_volume_Output_Profit_Loss'] * row['Carrier_loss_on_premium_Input_var']
            elif para == 'Working_capital_ow_Loss_Output_Cash_flow':
                row[para] = 0
                self.financial_report.loc[work_cap_timeline, para] = row['Costs_Output_Profit_Loss'] * row['Working_capital_ratio_carrier_loss_Input_var']
            elif para == 'Working_capital_ow_Distribution_channel_Output_Cash_flow':
                if self.current_month < self.start_month + ONE_YR:
                    self.financial_report.loc[work_cap_timeline, para] = row['ow_Distribution_channel_Output_Profit_Loss'] * row['Working_capital_ratio_Distribution_channel_Input_var']
                row[para] = 0
            elif para == 'Working_capital_ow_expenses_Output_Cash_flow':
                self.financial_report.loc[work_cap_timeline, para] = row['ow_expenses_Output_Profit_Loss'] * row['Working_capital_ratio_Expenses_Input_var']
                row[para] = 0
            elif para == 'Working_capital_Output_Cash_flow':
                wc1 = 'Working_capital_ow_Loss_Output_Cash_flow'
                wc2 = 'Working_capital_ow_Distribution_channel_Output_Cash_flow'
                wc3 = 'Working_capital_ow_expenses_Output_Cash_flow'
                row[para] = row[wc1] + row[wc2] + row[wc3]
                self.financial_report.loc[work_cap_timeline, para] = self.financial_report.loc[work_cap_timeline, wc1] + self.financial_report.loc[work_cap_timeline, wc2] + self.financial_report.loc[work_cap_timeline, wc3]
            elif para == 'Operating_cash_flow_Output_Cash_flow':
                nopat = 'NOPAT_Output_Cash_flow'
                wc = 'Working_capital_Output_Cash_flow'
                self.financial_report.loc[work_cap_timeline, para] = self.financial_report.loc[work_cap_timeline, nopat] - self.financial_report.loc[work_cap_timeline, wc]
                row[para] = row['NOPAT_Output_Cash_flow'] - row['Working_capital_Output_Cash_flow']
            else:
                pass
        return row

    def get_first_or_second_yr_ratio(self, ratio1, ratio2):
        ratio = 0
        if self.current_month - self.start_month < ONE_YR:
            ratio = ratio1
        else:
            ratio = ratio2
        return ratio

    def output_financial_report(self):
        return self.financial_report.copy()


###############
# Dataloader ##
###############
class DataLoader:
    '''
    To create list of output variables and list of dataframes.
    '''
    workflow_vars_sheet_name = 'Input_variables'
    output_vars_sheet_name = 'Output_variables_layout'
    def __init__(self, datapath, variable_path=None):
        self.filepath = datapath
        self.var_path = variable_path
        self.sheet_name_list = self.access_raw_sheetnames()

    ## access customer growth data
    def access_customer_growth_number(self):
        test = pd.read_excel(self.filepath, sheet_name='Customer growth', skiprows=1)
        test.index = test['Month']
        test_copy = test.transpose()
        test_new = test_copy.loc[0:480].copy()
        nn = test_new.to_numpy()
        new_list = [n[0] for n in nn]
        return new_list

    def access_raw_sheetnames(self):
        from openpyxl import load_workbook
        workbook = load_workbook(filename=self.filepath)
        return workbook.sheetnames

    def create_single_month_dict(self):
        single_value_df = [x for x in self.sheet_name_list if x.__contains__('input Cohort') and not x.__contains__('-')]
        single_val_list = [int(x.replace('input Cohort M', '')) for x in single_value_df]
        single_val_dic = {single_val_list[idx]: single_value_df[idx] for idx in range(len(single_value_df))}
        return single_val_dic

    def create_multiply_month_dict(self):
        multipl_values_df = [x for x in self.sheet_name_list if x.__contains__('input Cohort') and x.__contains__('-')]
        multi_val_dic = {}
        for idx in range(len(multipl_values_df)):
            temp = multipl_values_df[idx]
            z = temp.replace('input Cohort M', '')
            w = [int(x) for x in z.split(sep='-')]
            new_range = range(w[0], w[-1])
            multi_val_dic[new_range] = multipl_values_df[idx]
        return multi_val_dic


    def createDataFrameworkSuffix(self, df):
        df['Parameter'] = df['Parameter'] + '_' + df['Input/Output'] + '_' + df['Category']
        return df

    def output_variables(self, sheetname):
        data = pd.read_excel(self.var_path, sheet_name=sheetname, skiprows=1)
        df = self.createDataFrameworkSuffix(data)
        outputs = list(df.loc[df['Input/Output'] == 'Output', 'Parameter'])
        return outputs

    def extract_dataframe_from_sheet(self, sheetname):
        df = pd.read_excel(config.data_path, sheet_name=sheetname, skiprows=1)
        improved_df = self.createDataFrameworkSuffix(df)
        index_names = improved_df['Parameter']
        improved_df = improved_df.filter(like="Value", axis=1)
        improved_df.index = index_names
        improved_df.columns = improved_df.loc['Month_Input_var'].astype('int32')
        return improved_df.copy()


